from pathlib import Path
import openpyxl
import os
import xlrd
import json


form_json = {"TOTAL": None, "CATEGORY_V": None, "CATEGORY_IA": {"TOTAL": None, "LANES_4": None, "LANES_6": None, "LANES_8": None}, "CATEGORY_IB": {"TOTAL": None, "LANES_4": None, "LANES_6": None, "LANES_8": None}, "CATEGORY_IC": {"TOTAL": None, "LANES_4": None, "LANES_6": None, "LANES_8": None}, "CATEGORY_II": {"TOTAL": None, "LANES_4": None, "LANES_6": None, "LANES_2OR3": None}, "CATEGORY_IV": None, "CATEGORY_III": None}
params ='' #тут параметры формы отчетности

#парсинг значений таблицы
def parse_values_table(mas, region):
    ids = [[[None, None] for i in range(23)] for i in range(7)]
    a, b, c, d, e, f, k = -1, -1, -1, -1, -1, -1, -1
    if not mas:
        with open('not_checked.txt', "a") as file:
            file.write("incorrect sheet:        " + region + '\n')
        return ids
    max_str = len(mas)
    max_col = len(mas[0])
    for i in range(0, max_str):
        filtered_list = [item for item in mas[i] if item is not None and item != ""]
        if filtered_list == [1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0, 13.0, 14.0, 15.0, 16.0, 17.0, 18.0, 19.0, 20.0, 21.0, 22.0] or filtered_list == [1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0, 13.0, 14.0, 15.0, 16.0, 17.0, 18.0, 19.0, 20.0, 21.0, 22.0, 23.0]:
            for j in range(0, max_col):
                if mas[i][j] != '' and mas[i][j] is not None:
                    ids[0][int(mas[i][j]) - 1] = [i, j]
        if k == -1:
            for j in range(0, max_col):
                if (mas[i][j] == 1 or mas[i][j] == '1' or mas[i][j] == "01") and i > ids[0][0][0]:
                    k = j
                    ids[1][1] = [i, j]
        else:
            if (mas[i][k] == 2 or mas[i][k] == '2' or mas[i][k] == "02") and i > ids[0][0][0]:
                ids[2][1] = [i, k]
            if (mas[i][k] == 3 or mas[i][k] == '3' or mas[i][k] == "03") and i > ids[0][0][0]:
                ids[3][1] = [i, k]
            if (mas[i][k] == 4 or mas[i][k] == '4' or mas[i][k] == "04") and i > ids[0][0][0]:
                ids[4][1] = [i, k]
            if (mas[i][k] == 5 or mas[i][k] == '5' or mas[i][k] == "05") and i > ids[0][0][0]:
                ids[5][1] = [i, k]
            if (mas[i][k] == 6 or mas[i][k] == '6' or mas[i][k] == "06") and i > ids[0][0][0]:
                ids[6][1] = [i, k]
    x_cords = [elem[1] for elem in ids[0] if elem[1] is not None][2:]
    y_cords = [elem[1][0] for elem in ids if elem[1][0] is not None][1:]
    ans = [[None for i in range(21)] for j in range(5)]
    k, m = 0, 0
    for i in y_cords:
        m = 0
        for j in x_cords:
            if isinstance(mas[i][j], str) or mas[i][j] is None:
                ans[k][m] = None
            else:
                ans[k][m] = round(mas[i][j], 3)
            m += 1
        k += 1
    for i in range(len(ans)):
        if ans[i][0] == 'км':
            ans[i] = ans[i][1:]
        else:
            ans[i] = ans[i][:20]
    return ans

#парсинг значений колонок
def parse_value_col(mas, region):
    y_coords = [0] * 10
    ans = [[None for i in range(20)] for j in range(10)]
    if not mas:
        with open('not_checked.txt', "a") as file:
            file.write("incorrect sheet:        " + region + '\n')
        return ans
    f = 0
    for i in range(len(mas)):
        for j in range(len(mas[0])):
            if mas[i][j] == 5 or mas[i][j] == '5' or mas[i][j] == "05":
                if f == 1:
                    end = i
                f += 1
    k = 0
    for i in range(len(mas)):
        for j in range(end + 1, len(mas[0])):
            if isinstance(mas[i][j], str) and 'км' in mas[i][j]:
                y_coords[k] = i
                k += 1
    for i in range(len(y_coords)):
        f = False
        for elem in mas[y_coords[i]]:
            if isinstance(elem, int) or isinstance(elem, float):
                if not f:
                    ans[i][0] = round(elem, 3)
                f = True
        if not f:
            ans[i][0] = None
    return ans


def get_sheet_names(file_path):
    if file_path.lower().endswith('.xls'):
        # Для файлов .xls
        workbook = xlrd.open_workbook(file_path)
        sheet_names = workbook.sheet_names()

    elif file_path.lower().endswith('.xlsx'):
        # Для файлов .xlsx
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames

    else:
        print(f"Неподдерживаемый формат файла: {file_path}")
        return None

    return sheet_names


def check_correct_name(form_names, file_name):
    contains_any_element = any(item in file_name for item in form_names)
    if contains_any_element:
        return True
    return False


def excel_to_json():
    main_json = {}
    form_name = ["2-ДГ", "2ДГ", "2-DG", "2DG", "2-dg", "2dg", "2-дг", "2дг", "2 ДГ", "2 дг", "2 dg", "2 DG", "дг-2", "2- ДГ"]
    folder_path = '' #тут ссылка на путь к папаке с эксель файлами с формами отчетности
    files = os.listdir(folder_path)
    for elem in files:
        current_path = os.path.join(folder_path, elem)
        excels = os.listdir(current_path)

        #Проверка на 2-ДГ
        if len(excels) == 1:
            with open('not_checked.txt', "a") as file:
                file.write("universal form files:        " + elem + '\n')
            continue
        f = False
        for item in excels:
            if check_correct_name(form_name, item):
                if f:
                    f = False
                    break
                desired_file = item
                f = True
        if not f:
            with open('not_checked.txt', "a") as file:
                file.write("no or too many form files:   " + elem + '\n')
            continue

        current_path = os.path.join(current_path, desired_file)

        #Проверка на число листов
        is_problem = False
        if len(get_sheet_names(current_path)) > 3:
            is_problem = True

        if desired_file.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(current_path)
            second_sheet = workbook.worksheets[1]
            all_values = []
            for row in second_sheet.iter_rows(values_only=True):
                row_values = []
                for cell in row:
                    row_values.append(cell)
                all_values.append(row_values)
            if is_problem:
                main_json["?" + elem] = all_values
            else:
                main_json["+" + elem] = all_values
            workbook.close()

        if desired_file.endswith(".xls"):
            workbook = xlrd.open_workbook(current_path)
            sheet = workbook.sheet_by_index(1)
            all_values = []
            for row in range(sheet.nrows):
                row_values = []
                for col in range(sheet.ncols):
                    row_values.append(sheet.cell_value(row, col))
                all_values.append(row_values)
            if is_problem:
                main_json["?" + elem] = all_values
            else:
                main_json["+" + elem] = all_values
            workbook.release_resources()
    return main_json


if __name__ == '__main__':
    main_js = excel_to_json()

    log = False
    if log:
        for key, val in main_js.items():
            print(key)
            for elem in val:
                print(elem)

    for key, val in main_js.items():
        b = parse_values_table(val, key)
        a = parse_value_col(val, key)
        b.extend(a)
        main_js[key] = b

    for key, val in main_js.items():
        print(key)
        for elem in val:
            print(elem)

    with open('json_output.json', 'w', encoding='utf-8') as file:
        # Конвертация в JSON и запись в файл
        json.dump(main_js, file, ensure_ascii=False, indent=2)


